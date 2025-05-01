import streamlit as st
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
from scipy import stats
from statsmodels.tsa.arima.model import ARIMA
from datetime import datetime
from io import BytesIO
from fpdf import FPDF
from openpyxl import Workbook
from PIL import Image, ImageDraw, ImageOps
import plotly.graph_objects as go
import plotly.figure_factory as ff
import os
import logging
import tempfile

# --- Logging ---
logging.basicConfig(
    filename='app.log',
    level=logging.DEBUG,
    format='%(asctime)s - %(levelname)s - %(message)s'
)
# --- Data Validation ---
def validate_data_structure(data):
    return 'COMPANY' in data.columns

# --- Load CSV with validation ---
def load_data(file):
    try:
        data = pd.read_csv(file)
        if not validate_data_structure(data):
            st.error("Invalid structure! File must include a 'COMPANY' column.")
            return None
        data = data.fillna(0)
        data['COMPANY'] = data['COMPANY'].str.strip()
        return data
    except Exception as e:
        logging.error(f"Error loading data: {e}")
        st.error("There was an error processing the file. Please try again.")
        return None
# --- Generate a downloadable CSV template ---
def generate_template():
    template_df = pd.DataFrame({
        "COMPANY": ["ZAMBEEF", "ZANACO", "CEC"],
        "2021-01-01": [1.5, 2.1, 3.0],
        "2021-01-02": [1.6, 2.0, 3.1],
        "2021-01-03": [1.55, 2.05, 3.05]
    })
    buffer = BytesIO()
    template_df.to_csv(buffer, index=False)
    buffer.seek(0)
    return buffer
# --- Streamlit App ---
def app():
    st.markdown("""
    <style>
    body, .main, .stApp {
        background-color: black;
        color: white;
    }
    .stButton button {
        background-color: transparent !important;
        color: white !important;
        border: 1px solid white !important;
        padding: 10px 24px;
        font-size: 16px;
        border-radius: 5px;
        transition: 0.3s ease;
    }
    .stButton button:hover {
        background-color: red !important;
        color: white !important;
        border: 1px solid red !important;
    }
    </style>
""", unsafe_allow_html=True)

    st.title("Yengo | Lusaka Stock Exchange (LUSE) Portfolio Management")

    st.markdown("""
        About the Author: Limbani Phiri brings a wealth of experience from his extensive career across the banking, insurance, and commerce industries. As the CEO of Yengo, he has played a pivotal role in the development and creation of the Yengo platform.
    """)
    st.markdown(f"🕒 **{datetime.now().strftime('%A, %d %B %Y | %I:%M %p')}**")
    # --- Download Template Section ---
    st.markdown("### 📥 Download CSV Template")
    st.markdown("""
        To get started, download the CSV template below.  
        Fill it in with your own historical data, ensuring:
        - The first column is `COMPANY`
        - Subsequent columns are dates (format: YYYY-MM-DD)
        - Each row represents stock prices for a single company
    """)
    template_csv = generate_template()
    st.download_button(
        label="📄 Download Template CSV",
        data=template_csv,
        file_name="LUSE_template.csv",
        mime="text/csv"
    )
    uploaded_file = st.file_uploader("📁 Upload your LUSE Stock Data CSV file", type=["csv"])

    if uploaded_file:
        data = load_data(uploaded_file)
        if data is not None:
            st.markdown("### 🔍 Preview of Uploaded Data")
            st.dataframe(data.head())
            st.markdown("<hr class='red-line' />", unsafe_allow_html=True)

            try:
                temp = data.drop(columns=['COMPANY']).transpose()
                temp.columns = data['COMPANY'].values
                company_names = data['COMPANY'].values
                ModelData_Margins = np.diff(temp.values, axis=0)
                ModelData_Margins[ModelData_Margins == 0] = np.nan
            except Exception as e:
                logging.error(f"Error processing stock data: {e}")
                st.error("Could not process stock data. Check format.")
                return

            selected_company = st.selectbox("📊 Select a company for analysis", options=list(temp.columns))
            st.markdown(f"## 📊 Analysis for: `{selected_company}`")

            try:
                # --- Dynamic Chart Description ---
                st.markdown(f"### 📊 Price History for `{selected_company}`")
                st.markdown(f"This chart shows the historical closing prices for **{selected_company}** based on your uploaded data.")
                
                # --- Create Smooth Line Chart with Enhanced Styling ---
                fig = go.Figure()
                
                fig.add_trace(go.Scatter(
                    x=temp.index,
                    y=temp[selected_company],
                    mode='lines',
                    name='Closing Price',
                    line=dict(color='red', width=3, shape='spline'),  # Smooth curve
                    hovertemplate='Date: %{x}<br>Price: %{y:.2f}<extra></extra>'
                ))
                
                fig.update_layout(
                    title=f"{selected_company} Stock Price Over Time",
                    title_font=dict(size=22, color='red'),
                    xaxis_title="Date",
                    yaxis_title="Closing Price (ZMW)",
                    paper_bgcolor='black',
                    plot_bgcolor='black',
                    font=dict(color='red'),
                    xaxis=dict(showgrid=True, tickfont=dict(color='red')),
                    yaxis=dict(showgrid=True, tickfont=dict(color='red'))
                )
                
                st.plotly_chart(fig, use_container_width=True)
                # --- ARIMA Forecast ---
                # --- Dynamic Forecast Description ---
                st.markdown(f"### 🔮 300-Day ARIMA Forecast for `{selected_company}`")
                st.markdown(f"This chart displays a 300-day forecast of **{selected_company}**’s stock price using an ARIMA(5,1,0) model based on historical data.")
                
                # --- Fit ARIMA Model and Generate Forecast ---
                model = ARIMA(temp[selected_company], order=(5, 1, 0))
                fit = model.fit()
                forecast = fit.forecast(steps=300)
                
                # --- Create Forecast Plot with Enhanced Styling ---
                fig2 = go.Figure()
                fig2.add_trace(go.Scatter(
                    x=list(range(len(forecast))),
                    y=forecast,
                    mode='lines',
                    name='Forecast',
                    line=dict(color='red', width=3, shape='spline'),  # Smooth curve
                    hovertemplate='Day: %{x}<br>Forecasted Price: %{y:.2f}<extra></extra>'
                ))
                
                fig2.update_layout(
                    title=f"{selected_company} - 300-Day Forecast (ARIMA)",
                    title_font=dict(size=22, color='red'),
                    xaxis_title="Day Index (Future)",
                    yaxis_title="Forecasted Price (ZMW)",
                    paper_bgcolor='black',
                    plot_bgcolor='black',
                    font=dict(color='red'),
                    xaxis=dict(showgrid=True, tickfont=dict(color='red')),
                    yaxis=dict(showgrid=True, tickfont=dict(color='red'))
                )
                
                st.plotly_chart(fig2, use_container_width=True)


            except Exception as e:
                logging.error(f"Error in forecast or plot: {e}")
                st.error("There was a problem with forecasting.")
                return

            try:
                # --- Portfolio Simulation ---
                initial_portfolio = 1000
                initial_price = temp[selected_company].iloc[-1]
                initial_value = initial_portfolio * initial_price
                mean_margin = np.nanmean(ModelData_Margins[:, temp.columns.get_loc(selected_company)])
                std_margin = np.nanstd(ModelData_Margins[:, temp.columns.get_loc(selected_company)])

                simulated_prices = []
                for _ in range(10000):
                    returns = np.random.normal(loc=mean_margin, scale=std_margin, size=300)
                    path = initial_price + np.cumsum(returns)
                    simulated_prices.append(path[-1])

                # --- Dynamic Simulation Description ---
                st.markdown(f"### 🔮 Monte Carlo Simulation for `{selected_company}`")
                st.markdown(f"This chart shows the simulated distribution of **{selected_company}**’s stock price based on historical volatility and mean return.")
                
                # --- Create Simulated Distribution Plot with Enhanced Styling ---
                fig3 = ff.create_distplot([simulated_prices], group_labels=[f"{selected_company} Simulation"], show_hist=False, colors=['red'])
                
                fig3.update_layout(
                    title=f"{selected_company} - Simulated Distribution",
                    title_font=dict(size=22, color='red'),
                    xaxis_title="Simulated Price (ZMW)",
                    yaxis_title="Density",
                    paper_bgcolor='black',
                    plot_bgcolor='black',
                    font=dict(color='red'),
                    xaxis=dict(showgrid=True, tickfont=dict(color='red')),
                    yaxis=dict(showgrid=True, tickfont=dict(color='red'))
                )
                
                st.plotly_chart(fig3, use_container_width=True)


                # --- Summary ---
                expected_price = np.mean(simulated_prices)
                final_value = expected_price * initial_portfolio
                change = final_value - initial_value

                st.markdown("### 💼 Portfolio Prediction Summary")
                st.write(f"**Expected Final Portfolio Value**: K`{final_value:.2f}`")
                st.write(f"**Change in Value**: K`{change:.2f}`")
                from PIL import Image
                # --- Load and process logo ---
                logo_path = "Yengo_1.jpg"
                image = Image.open(logo_path).convert("L").convert("RGBA")
                min_side = min(image.size)
                image = image.crop((
                    (image.width - min_side) // 2,
                    (image.height - min_side) // 2,
                    (image.width + min_side) // 2,
                    (image.height + min_side) // 2
                ))
                
                mask = Image.new('L', image.size, 0)
                draw = ImageDraw.Draw(mask)
                draw.ellipse((0, 0, image.size[0], image.size[1]), fill=255)
                image.putalpha(mask)
                
                border_size = 8
                border_image = Image.new(
                    'RGBA',
                    (image.size[0] + border_size*2, image.size[1] + border_size*2),
                    (255, 255, 255, 0)
                )
                border_mask = Image.new('L', border_image.size, 0)
                border_draw = ImageDraw.Draw(border_mask)
                border_draw.ellipse(
                    (0, 0, border_image.size[0], border_image.size[1]),
                    fill=255
                )
                border_image.paste(image, (border_size, border_size), image)
                border_image.putalpha(border_mask)
                # Use the processed circular logo image
                logo_image = border_image  
                # --- Generate PDF buffer ---
                pdf_buffer = export_pdf_summary(
                    temp=temp,
                    forecast=forecast,
                    closing_prices=simulated_prices,
                    selected_company=selected_company,
                    logo_image=logo_image
                )
                
                # --- Download Button ---
                st.markdown("<hr class='red-line' />", unsafe_allow_html=True)
                
                if 'pdf_buffer' in locals() and pdf_buffer:
                    st.markdown("### 📤 Download Your Forecast Report")
                    st.info("Click the button below to download your personalized LUSE stock forecast summary as a PDF.")
                    
                    st.download_button(
                        label="📄 Download PDF Report",
                        data=pdf_buffer,
                        file_name=f"{selected_company}_LUSE_Report.pdf",
                        mime="application/pdf"
                    )



            except Exception as e:
                logging.error(f"Simulation error: {e}")
                st.error("There was a problem simulating the portfolio.")
                return
        else:
            st.warning("Please upload a valid CSV file.")
    else:
        st.warning("🚨 Please upload a CSV with `COMPANY` column and historical prices.")

# --- PDF Export ---
def export_pdf_summary(temp, forecast, closing_prices, selected_company, logo_image):
    try:
        pdf = FPDF()
        pdf.add_page()
        pdf.set_auto_page_break(auto=True, margin=15)

        # Save logo to a temporary file
        with tempfile.NamedTemporaryFile(delete=False, suffix=".png") as tmp:
            logo_image.save(tmp.name, format="PNG")
            logo_path = tmp.name

        # Add logo to PDF
        pdf.set_font("Arial", 'B', 14)
        pdf.image(logo_path, x=10, y=8, w=30)
        pdf.cell(0, 10, f"Report Number: {np.random.randint(1000, 9999)}", align='R', ln=1)
        pdf.cell(0, 10, datetime.now().strftime('%A, %d %B %Y | %I:%M %p'), align='C', ln=1)
        pdf.ln(5)

        # Title and company info
        pdf.set_font("Arial", 'B', 18)
        pdf.set_text_color(220, 50, 50)
        pdf.cell(0, 15, "Lusaka Stock Exchange Portfolio Prediction", ln=1, align='C')

        pdf.set_font("Arial", 'B', 14)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(0, 10, f"Company: {selected_company}", ln=1, align='C')

        # Summary Table
        col_index = list(temp.columns).index(selected_company)
        last_price = temp.iloc[-1, col_index]
        expected_price = np.mean(closing_prices)
        final_value = expected_price * 1000
        change = final_value - (last_price * 1000)

        pdf.set_font("Arial", 'B', 12)
        for label, value, analysis in [
            ("Last Known Price", f"K{last_price:.2f}", "-"),
            ("Forecasted Avg. Price", f"K{expected_price:.2f}", "-"),
            ("Portfolio Value (x1000)", f"K{final_value:.2f}", "Positive" if change >= 0 else "Negative"),
            ("Expected Change", f"K{change:.2f}", "Positive" if change >= 0 else "Negative"),
        ]:
            pdf.cell(70, 10, label, 1)
            pdf.cell(60, 10, value, 1)
            pdf.cell(50, 10, analysis, 1, 1)

        # Footer
        pdf.set_y(-30)
        pdf.set_font("Arial", 'B', 10)
        pdf.cell(0, 10, "DISCLAIMER: This is a simulation and does not guarantee future performance.", 0, 1, 'C')
        pdf.cell(0, 10, "© 2025 Yengo. All rights reserved.", 0, 1, 'C')

        # Write PDF to BytesIO using 'S' mode
        pdf_output = pdf.output(dest='S').encode('latin1')
        buffer = BytesIO(pdf_output)
        return buffer

    except Exception as e:
        logging.error(f"PDF export failed: {e}")
        st.error("Failed to generate PDF summary.")
        return None


# --- Main Execution ---
if __name__ == "__main__":
    app
